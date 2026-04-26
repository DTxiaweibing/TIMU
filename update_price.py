import json
import re

try:
    import pandas as pd
    ENGINE = "pandas"
except ImportError:
    from openpyxl import load_workbook
    ENGINE = "openpyxl"

EXCEL_FILE = "price.xlsx"
JSON_FILE = "price.json"

def is_lobster(name):
    lobsters = ['龙虾', '澳洲龙虾', '波士顿龙虾', '小青龙', '花龙', '青龙']
    return any(lob in name for lob in lobsters)

def is_shrimp(name):
    """判断是否是虾类（需要按大（贵）→小（便宜）分配）"""
    shrimps = ['虾']
    return any(s in name for s in shrimps)

def is_fish(name):
    """判断是否是鱼类（大鱼便宜，小鱼贵）"""
    fishes = ['鱼']
    return any(f in name for f in fishes)

def process_item(name, price, target_list):
    original_name = name.strip()
    cleaned_name = original_name.replace('（', '(').replace('）', ')').strip()
    
    # 河虾、龙虾不拆分，直接保留
    if '河虾' in cleaned_name or is_lobster(cleaned_name):
        target_list.append({"name": original_name, "price": price})
        return

    prices = re.split(r'[-~—]', price)
    prices = [p.strip() for p in prices if p.strip()]
    price_count = len(prices)

    if price_count == 1:
        target_list.append({"name": original_name, "price": prices[0]})
        return

    base_name = re.sub(r'\([^)]*\)', '', cleaned_name).strip()

    if price_count == 2:
        # 比较价格大小，决定哪个是大哪个是小
        try:
            p1 = float(prices[0])
            p2 = float(prices[1])
        except ValueError:
            # 解析不了就保留原样
            target_list.append({"name": original_name, "price": price})
            return

        if p1 == p2:
            specs = ['大', '小']
            ordered_prices = prices
        elif is_fish(cleaned_name):
            # 鱼：大鱼便宜，小鱼贵，所以价格低的为大
            if p1 < p2:
                specs = ['大', '小']
                ordered_prices = prices
            else:
                specs = ['小', '大']
                ordered_prices = [prices[1], prices[0]]
        else:
            # 虾或其他：大虾贵，小虾便宜，所以价格高的为大
            if p1 > p2:
                specs = ['大', '小']
                ordered_prices = prices
            else:
                specs = ['小', '大']
                ordered_prices = [prices[1], prices[0]]

        for i, spec in enumerate(specs):
            target_list.append({
                "name": f"{base_name}（{spec}）",
                "price": ordered_prices[i]
            })
        return

    if price_count == 3:
        # 三个价格：鱼类按价格升序（低-中-高）→大-中-小，虾类按降序（高-中-低）→大-中-小
        try:
            p_vals = [float(p) for p in prices]
        except ValueError:
            target_list.append({"name": original_name, "price": price})
            return

        if is_fish(cleaned_name):
            # 鱼：低=大，中=中，高=小
            sorted_indices = sorted(range(3), key=lambda i: p_vals[i])  # 升序
            spec_map = {sorted_indices[0]: '大', sorted_indices[1]: '中', sorted_indices[2]: '小'}
        else:
            # 虾：高=大，中=中，低=小
            sorted_indices = sorted(range(3), key=lambda i: p_vals[i], reverse=True)  # 降序
            spec_map = {sorted_indices[0]: '大', sorted_indices[1]: '中', sorted_indices[2]: '小'}

        specs = [spec_map[i] for i in range(3)]
        for i, spec in enumerate(specs):
            target_list.append({
                "name": f"{base_name}（{spec}）",
                "price": prices[i]
            })
        return

    # 超过3个价格，保留原样
    target_list.append({"name": original_name, "price": price})

def process_lobster(row, target_list, last_price):
    if len(row) < 4:
        return last_price
    name = row[1].strip() if row[1] else ''
    spec = row[2].strip() if row[2] else ''
    price_str = row[3].strip() if row[3] else ''
    if '龙虾' not in name or not spec:
        return last_price
    if not price_str:
        price_str = last_price
    else:
        price_str = price_str.replace('-', '~')
        last_price = price_str
    if not price_str:
        return last_price
    target_list.append({"name": f"龙虾（{spec}）", "price": price_str})
    return last_price

def normalize_date(raw_date):
    if not raw_date:
        return ""
    raw_date = str(raw_date).strip()
    match = re.search(r'(\d{4})[./-](\d{1,2})[./-](\d{1,2})', raw_date)
    if match:
        year, month, day = match.groups()
        return f"{year}/{int(month):02d}/{int(day):02d}"
    return raw_date

def parse_first_price(price_str):
    if not price_str:
        return 0.0
    parts = re.split(r'[~\-]', price_str)
    for part in parts:
        try:
            return float(part.strip())
        except ValueError:
            continue
    return 0.0

def extract_base_name(item_name):
    """提取商品的基础名称，用于分组"""
    name = item_name.strip()
    # 去掉末尾的（大）（中）（小）
    name = re.sub(r'[（(][大中小][）)]$', '', name).strip()
    # 对于河虾，去掉头数，如(100头)、(130头)
    name = re.sub(r'\(\d+头\)$', '', name).strip()
    return name

def sort_items(items):
    """组内排序：河虾组整体放最后，内部按价格降序；
       其他虾按大中小排列；鱼类按大中小（价格已在上一步修正）"""
    # 分组并记录首次出现顺序
    grouped = {}
    group_order = []
    for item in items:
        base = extract_base_name(item["name"])
        if base not in grouped:
            grouped[base] = []
            group_order.append(base)
        grouped[base].append(item)

    # 组内排序
    spec_order = {'大': 0, '中': 1, '小': 2}
    for base in grouped:
        lst = grouped[base]
        if '河虾' in base:
            lst.sort(key=lambda x: parse_first_price(x["price"]), reverse=True)
        else:
            # 按大中小的索引排序
            def spec_key(item):
                name = item["name"]
                match = re.search(r'[（(]([大中小])[）)]$', name)
                if match:
                    return spec_order.get(match.group(1), 99)
                return 99
            lst.sort(key=spec_key)

    # 河虾组移动到末尾
    if '河虾' in group_order:
        group_order.remove('河虾')
        group_order.append('河虾')

    result = []
    for base in group_order:
        result.extend(grouped[base])
    return result

def main():
    if ENGINE == "pandas":
        df = pd.read_excel(EXCEL_FILE, header=None, dtype=str)
        rows = df.values.tolist()
    else:
        wb = load_workbook(EXCEL_FILE, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            rows.append([str(c).strip() if c is not None else '' for c in row])
        wb.close()

    categories = []
    date_str = ""
    current_cat = None
    in_lobster = False
    lobster_last_price = ""

    for row in rows:
        if not row or all((pd.isna(c) if ENGINE == "pandas" else c == '') for c in row):
            continue
        row = [str(c).strip() if (not (pd.isna(c) if ENGINE == "pandas" else c == '')) else '' for c in row]
        row_text = ' '.join(row)

        if '日期:' in row_text:
            parts = row_text.split('日期:')
            if len(parts) > 1:
                date_str = normalize_date(parts[1])
            continue

        first_cell = row[0]
        if '批发价' in first_cell:
            title = first_cell.replace('：', ':').split(':')[0].strip()
            current_cat = {"title": title, "items": []}
            categories.append(current_cat)
            if '龙虾' in title:
                in_lobster = True
                lobster_last_price = ""
            else:
                in_lobster = False
            continue

        if first_cell in ['编号', '品名']:
            continue

        if current_cat is not None:
            if in_lobster:
                lobster_last_price = process_lobster(row, current_cat["items"], lobster_last_price)
            else:
                if len(row) >= 3 and row[1] and row[2]:
                    process_item(row[1], row[2], current_cat["items"])
                if len(row) >= 6 and row[4] and row[5]:
                    process_item(row[4], row[5], current_cat["items"])

    for cat in categories:
        cat["items"] = sort_items(cat["items"])

    result = {"date": date_str, "categories": categories}
    with open(JSON_FILE, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    total_items = sum(len(c['items']) for c in categories)
    print(f"✅ 成功生成 {JSON_FILE}，共 {len(categories)} 个分类，{total_items} 条记录")

if __name__ == "__main__":
    main()